import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill

def export_to_excel(queryset, filename, headers):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
    # Write data
    for row_num, obj in enumerate(queryset, 2):
        for col_num, (field, field_type) in enumerate(headers.items(), 1):
            val = getattr(obj, field)
            # Handle Foreign Keys
            if hasattr(val, '__str__') and not isinstance(val, (str, int, float, bool)):
                val = str(val)
            ws.cell(row=row_num, column=col_num, value=val)
            
    # Auto-adjust column width
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
        
    wb.save(response)
    return response
